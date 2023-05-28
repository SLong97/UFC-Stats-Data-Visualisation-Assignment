from collections import ChainMap

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
import pandas as pd
import time


book = load_workbook('UFC_STATS.xlsx')
writer = pd.ExcelWriter('UFC_STATS.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

serve = Service("C:\\Users\\Seán Long\\Downloads\\geckodriver.exe")
driver = webdriver.Firefox(service=serve)
driver.get('http://www.ufcstats.com/statistics/events/completed')

ufc_events = driver.find_elements(by=By.CSS_SELECTOR, value='.b-link.b-link_style_black')

events = []
for event in ufc_events:
    eventURL = event.get_attribute("href")
    eventName = event.text
    events.append({"event_name": eventName, "event_url": eventURL})

fights = []
for event in events:
    driver.get(event["event_url"])
    fightsElements = driver.find_elements(by=By.CSS_SELECTOR, value='.b-fight-details__table-row.b-fight-details__table-row__hover.js-fight-details-click')

    for element, division in enumerate(fightsElements):

        fightURL = fightsElements[element].get_attribute("data-link")
        division_name = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col l-page_align_left"]/p')[2].text
        fighter1 = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col l-page_align_left"]/p')[0].text
        fighter2 = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col l-page_align_left"]/p')[1].text
        bout = fighter1 + " vs. " + fighter2
        victory_method = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col l-page_align_left"]/p')[3].text
        final_round = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col"]/p')[8].text
        round_time = division.find_elements(by=By.XPATH, value='.//td[@class="b-fight-details__table-col"]/p')[9].text


        fights.append({"event_name" : event["event_name"], "fight_url" : fightURL, "division_name" : division_name, "bout" : bout,
                       "winner": fighter1, "victory_method": victory_method, "final_round" : final_round, "round_time" : round_time})

#fights = [{"fight_url" : "http://www.ufcstats.com/fight-details/0ccd2593c88209e4"}]
#stats = []
s = 0
final_df = pd.DataFrame()
for fight in fights:

    stats = []

    driver.get(fight["fight_url"])

    round_by_round_stats = driver.find_elements(by=By.CSS_SELECTOR, value='a.b-fight-details__collapse-link_rnd.js-fight-collapse-link')
    # All Totals Table
    round_by_round_stats[0].click()
    # All Significant Strikes Table
    round_by_round_stats[1].click()

    # FIGHTER NAMES LINK

    nameElement = driver.find_elements(by=By.XPATH, value='//a[@class="b-link b-fight-details__person-link"]')

    detailsElement = driver.find_elements(by=By.XPATH, value='//i[@class="b-fight-details__text-item"]')

    # TOTALS TABLE
    table_body = driver.find_elements(By.CLASS_NAME, "b-fight-details__table-body")[0]
    elements = table_body.find_elements(By.CSS_SELECTOR, "td")

    # ROUND BY ROUND
    RBR_table_body = driver.find_elements(By.CSS_SELECTOR, ".b-fight-details__table.js-fight-table")[0]
    RBR_elements = RBR_table_body.find_elements(By.CLASS_NAME, "b-fight-details__table-row")


    # SIGNIFICANT STRIKES TOTALS TABLE
    table_body2 = driver.find_elements(By.CLASS_NAME, "b-fight-details__table-body")[2]
    elements2 = table_body2.find_elements(By.CSS_SELECTOR, "td")

    # ROUND BY ROUND
    RBR_table_body2 = driver.find_elements(By.CSS_SELECTOR, ".b-fight-details__table.js-fight-table")[1]
    RBR_elements2 = RBR_table_body2.find_elements(By.CLASS_NAME, "b-fight-details__table-row")


             # FIGHTER NAMES

    red_fighter = nameElement[0].text
    blue_fighter = nameElement[1].text

    fight_format = detailsElement[2].text.split(" ")[2]

    stats.append({"EVENT" : fight["event_name"],
                  "BOUT" : fight["bout"],
                  "DIVISION": fight["division_name"],
                  "ROUNDS" : fight_format,
                  "WINNER": fight["winner"],
                  "METHOD OF WIN" : fight["victory_method"],
                  "FINAL ROUND": fight["final_round"],
                  "TIME" : fight["round_time"],
                  "RED_FIGHTER" : red_fighter,
                  "BLUE_FIGHTER" : blue_fighter})



            # FIGHTER TOTALS

    # Knockdowns

    red_fighter_kd = elements[1].text.split("\n")[0]
    blue_fighter_kd = elements[1].text.split("\n")[1]

    # Significant Strikes

    fighter_ss_count = elements[2].text
    red_fighter_ss_landed = fighter_ss_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_attempted = fighter_ss_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_landed = fighter_ss_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_attempted = fighter_ss_count.split("\n")[1].split(" of ")[1]

    # Significant Strikes Percentage

    fighter_ss_percent = elements[3].text
    red_fighter_ss_percent = fighter_ss_percent.split("\n")[0]
    blue_fighter_ss_percent = fighter_ss_percent.split("\n")[1]

    # Total Strikes

    fighter_ts_count = elements[4].text
    red_fighter_ts_landed = fighter_ts_count.split("\n")[0].split(" of ")[0]
    red_fighter_ts_attempted = fighter_ts_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ts_landed = fighter_ts_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ts_attempted = fighter_ts_count.split("\n")[1].split(" of ")[1]

    # Take-downs

    fighter_td_count = elements[5].text
    red_fighter_td_landed = fighter_td_count.split("\n")[0].split(" of ")[0]
    red_fighter_td_attempted = fighter_td_count.split("\n")[0].split(" of ")[1]
    blue_fighter_td_landed = fighter_td_count.split("\n")[1].split(" of ")[0]
    blue_fighter_td_attempted = fighter_td_count.split("\n")[1].split(" of ")[1]

    # Take-downs Percentage

    fighter_td_percent = elements[6].text
    red_fighter_td_percent = fighter_td_percent.split("\n")[0]
    blue_fighter_td_percent = fighter_td_percent.split("\n")[1]

    # Submission Attempts

    fighter_sub_attempts = elements[7].text
    red_fighter_sub_attempts = fighter_sub_attempts.split("\n")[0]
    blue_fighter_sub_attempts = fighter_sub_attempts.split("\n")[1]

    # Reversals

    fighter_reversals = elements[8].text
    red_fighter_reversals = fighter_reversals.split("\n")[0]
    blue_fighter_reversals = fighter_reversals.split("\n")[1]

    # Control Time

    fighter_control_time = elements[9].text
    red_fighter_control_time = fighter_control_time.split("\n")[0]
    blue_fighter_control_time = fighter_control_time.split("\n")[1]

    stats.append({"RED_FIGHTER_KD" : red_fighter_kd,
                  "RED_FIGHTER_SS_LANDED" : red_fighter_ss_landed,
                  "RED_FIGHTER_SS_ATTEMPTED" : red_fighter_ss_attempted,
                  "RED_FIGHTER_SS_PERCENT" : red_fighter_ss_percent,
                  "RED_FIGHTER_TS_LANDED" : red_fighter_ts_landed,
                  "RED_FIGHTER_TS_ATTEMPTED" : red_fighter_ts_attempted,
                  "RED_FIGHTER_TD_LANDED" : red_fighter_td_landed,
                  "RED_FIGHTER_TD_ATTEMPTED" : red_fighter_td_attempted,
                  "RED_FIGHTER_TD_PERCENT" : red_fighter_td_percent,
                  "RED_FIGHTER_SUB_ATTEMPTS" : red_fighter_sub_attempts,
                  "RED_FIGHTER_REVERSALS" : red_fighter_reversals,
                  "RED_FIGHTER_CONTROL_TIME" : red_fighter_control_time,

                  "BLUE_FIGHTER_KD": blue_fighter_kd,
                  "BLUE_FIGHTER_SS_LANDED": blue_fighter_ss_landed,
                  "BLUE_FIGHTER_SS_ATTEMPTED": blue_fighter_ss_attempted,
                  "BLUE_FIGHTER_SS_PERCENT": blue_fighter_ss_percent,
                  "BLUE_FIGHTER_TS_LANDED": blue_fighter_ts_landed,
                  "BLUE_FIGHTER_TS_ATTEMPTED": blue_fighter_ts_attempted,
                  "BLUE_FIGHTER_TD_LANDED": blue_fighter_td_landed,
                  "BLUE_FIGHTER_TD_ATTEMPTED": blue_fighter_td_attempted,
                  "BLUE_FIGHTER_TD_PERCENT": blue_fighter_td_percent,
                  "BLUE_FIGHTER_SUB_ATTEMPTS": blue_fighter_sub_attempts,
                  "BLUE_FIGHTER_REVERSALS": blue_fighter_reversals,
                  "BLUE_FIGHTER_CONTROL_TIME": blue_fighter_control_time
                  })



            # FIGHTER TOTALS ROUND BY ROUND

    for i in range(2,len(RBR_elements),2):

        round = ("_R" + str(int(i / 2)))
        #print(round)

        # Knockdowns

        red_fighter_kd_r = RBR_elements[i].text.split("\n")[2]
        blue_fighter_kd_r = RBR_elements[i].text.split("\n")[3]


        # Significant Strikes

        fighter_ss_count = RBR_elements[i].text
        red_fighter_ss_landed_r = fighter_ss_count.split("\n")[4].split(" of ")[0]
        red_fighter_ss_attempted_r = fighter_ss_count.split("\n")[4].split(" of ")[1]
        blue_fighter_ss_landed_r = fighter_ss_count.split("\n")[5].split(" of ")[0]
        blue_fighter_ss_attempted_r = fighter_ss_count.split("\n")[5].split(" of ")[1]


        # Significant Strikes Percentage

        fighter_ss_percent = RBR_elements[i].text
        red_fighter_ss_percent_r = fighter_ss_percent.split("\n")[6]
        blue_fighter_ss_percent_r = fighter_ss_percent.split("\n")[7]


        # Total Strikes

        fighter_ts_count = RBR_elements[i].text
        red_fighter_ts_landed_r = fighter_ts_count.split("\n")[8].split(" of ")[0]
        red_fighter_ts_attempted_r = fighter_ts_count.split("\n")[8].split(" of ")[1]
        blue_fighter_ts_landed_r = fighter_ts_count.split("\n")[9].split(" of ")[0]
        blue_fighter_ts_attempted_r = fighter_ts_count.split("\n")[9].split(" of ")[1]


        # Take-downs

        fighter_td_count = RBR_elements[i].text
        red_fighter_td_landed_r = fighter_td_count.split("\n")[10].split(" of ")[0]
        red_fighter_td_attempted_r = fighter_td_count.split("\n")[10].split(" of ")[1]
        blue_fighter_td_landed_r = fighter_td_count.split("\n")[11].split(" of ")[0]
        blue_fighter_td_attempted_r = fighter_td_count.split("\n")[11].split(" of ")[1]


        # Take-downs Percentage

        fighter_td_percent = RBR_elements[i].text
        red_fighter_td_percent_r = fighter_td_percent.split("\n")[12]
        blue_fighter_td_percent_r = fighter_td_percent.split("\n")[13]


        # Submission Attempts

        fighter_sub_attempts = RBR_elements[i].text
        red_fighter_sub_attempts_r = fighter_sub_attempts.split("\n")[14]
        blue_fighter_sub_attempts_r = fighter_sub_attempts.split("\n")[15]


        # Reversals

        fighter_reversals = RBR_elements[i].text
        red_fighter_reversals_r = fighter_reversals.split("\n")[16]
        blue_fighter_reversals_r = fighter_reversals.split("\n")[17]


        # Control Time

        fighter_control_time = RBR_elements[i].text
        red_fighter_control_time_r = fighter_control_time.split("\n")[18]
        blue_fighter_control_time_r = fighter_control_time.split("\n")[19]

        stats.append({"RED_FIGHTER_KD"+round : red_fighter_kd_r,
                      "RED_FIGHTER_SS_LANDED"+round : red_fighter_ss_landed_r,
                      "RED_FIGHTER_SS_ATTEMPTED"+round : red_fighter_ss_attempted_r,
                      "RED_FIGHTER_SS_PERCENT"+round : red_fighter_ss_percent_r,
                      "RED_FIGHTER_TS_LANDED"+round : red_fighter_ts_landed_r,
                      "RED_FIGHTER_TS_ATTEMPTED"+round : red_fighter_ts_attempted_r,
                      "RED_FIGHTER_TD_LANDED"+round : red_fighter_td_landed_r,
                      "RED_FIGHTER_TD_ATTEMPTED"+round : red_fighter_td_attempted_r,
                      "RED_FIGHTER_TD_PERCENT"+round : red_fighter_td_percent_r,
                      "RED_FIGHTER_SUB_ATTEMPTS"+round : red_fighter_sub_attempts_r,
                      "RED_FIGHTER_REVERSALS"+round : red_fighter_reversals_r,
                      "RED_FIGHTER_CONTROL_TIME"+round : red_fighter_control_time_r,

                      "BLUE_FIGHTER_KD"+round : blue_fighter_kd_r,
                      "BLUE_FIGHTER_SS_LANDED"+round: blue_fighter_ss_landed_r,
                      "BLUE_FIGHTER_SS_ATTEMPTED"+round: blue_fighter_ss_attempted_r,
                      "BLUE_FIGHTER_SS_PERCENT"+round: blue_fighter_ss_percent_r,
                      "BLUE_FIGHTER_TS_LANDED"+round: blue_fighter_ts_landed_r,
                      "BLUE_FIGHTER_TS_ATTEMPTED"+round: blue_fighter_ts_attempted_r,
                      "BLUE_FIGHTER_TD_LANDED"+round: blue_fighter_td_landed_r,
                      "BLUE_FIGHTER_TD_ATTEMPTED"+round: blue_fighter_td_attempted_r,
                      "BLUE_FIGHTER_TD_PERCENT"+round: blue_fighter_td_percent_r,
                      "BLUE_FIGHTER_SUB_ATTEMPTS"+round: blue_fighter_sub_attempts_r,
                      "BLUE_FIGHTER_REVERSALS"+round: blue_fighter_reversals_r,
                      "BLUE_FIGHTER_CONTROL_TIME"+round: blue_fighter_control_time_r})

    # counter = 0
    # for stat in stats:
    #     for key, value in stat.items():
    #         print(key, ":", value)
    #         counter += 1
    #         if counter % 12 == 0:
    #             print("\n")



            # SIGNIFICANT STRIKES TOTALS

    # Significant Strikes

    fighter_ss2_count = elements2[1].text
    red_fighter_ss2_landed = fighter_ss2_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss2_attempted = fighter_ss2_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss2_landed = fighter_ss2_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss2_attempted = fighter_ss2_count.split("\n")[1].split(" of ")[1]

    # Significant Strikes Percentage

    fighter_ss2_percent = elements2[2].text
    red_fighter_ss2_percent = fighter_ss2_percent.split("\n")[0]
    blue_fighter_ss2_percent = fighter_ss2_percent.split("\n")[1]

    # Head

    fighter_ss_head_count = elements2[3].text
    red_fighter_ss_head_landed = fighter_ss_head_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_head_attempted = fighter_ss_head_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_head_landed = fighter_ss_head_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_head_attempted = fighter_ss_head_count.split("\n")[1].split(" of ")[1]

    # Body

    fighter_ss_body_count = elements2[4].text
    red_fighter_ss_body_landed = fighter_ss_body_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_body_attempted = fighter_ss_body_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_body_landed = fighter_ss_body_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_body_attempted = fighter_ss_body_count.split("\n")[1].split(" of ")[1]

    # Leg

    fighter_ss_leg_count = elements2[5].text
    red_fighter_ss_leg_landed = fighter_ss_leg_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_leg_attempted = fighter_ss_leg_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_leg_landed = fighter_ss_leg_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_leg_attempted = fighter_ss_leg_count.split("\n")[1].split(" of ")[1]

    # Distance

    fighter_ss_distance_count = elements2[6].text
    red_fighter_ss_distance_landed = fighter_ss_distance_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_distance_attempted = fighter_ss_distance_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_distance_landed = fighter_ss_distance_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_distance_attempted = fighter_ss_distance_count.split("\n")[1].split(" of ")[1]

    # Clinch

    fighter_ss_clinch_count = elements2[7].text
    red_fighter_ss_clinch_landed = fighter_ss_clinch_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_clinch_attempted = fighter_ss_clinch_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_clinch_landed = fighter_ss_clinch_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_clinch_attempted = fighter_ss_clinch_count.split("\n")[1].split(" of ")[1]

    # Ground

    fighter_ss_ground_count = elements2[8].text
    red_fighter_ss_ground_landed = fighter_ss_ground_count.split("\n")[0].split(" of ")[0]
    red_fighter_ss_ground_attempted = fighter_ss_ground_count.split("\n")[0].split(" of ")[1]
    blue_fighter_ss_ground_landed = fighter_ss_ground_count.split("\n")[1].split(" of ")[0]
    blue_fighter_ss_ground_attempted = fighter_ss_ground_count.split("\n")[1].split(" of ")[1]

    stats.append({"RED_FIGHTER_SS2_LANDED" : red_fighter_ss2_landed,
                  "RED_FIGHTER_SS2_ATTEMPTED" : red_fighter_ss2_attempted,
                  "RED_FIGHTER_SS2_PERCENT" : red_fighter_ss2_percent,
                  "RED_FIGHTER_SS_HEAD_LANDED" : red_fighter_ss_head_landed,
                  "RED_FIGHTER_SS_HEAD_ATTEMPTED" : red_fighter_ss_head_attempted,
                  "RED_FIGHTER_SS_BODY_LANDED" : red_fighter_ss_body_landed,
                  "RED_FIGHTER_SS_BODY_ATTEMPTED" : red_fighter_ss_body_attempted,
                  "RED_FIGHTER_SS_LEG_LANDED" : red_fighter_ss_leg_landed,
                  "RED_FIGHTER_SS_LEG_ATTEMPTED" : red_fighter_ss_leg_attempted,
                  "RED_FIGHTER_SS_DISTANCE_LANDED" : red_fighter_ss_distance_landed,
                  "RED_FIGHTER_SS_DISTANCE_ATTEMPTED" : red_fighter_ss_distance_attempted,
                  "RED_FIGHTER_SS_CLINCH_LANDED" : red_fighter_ss_clinch_landed,
                  "RED_FIGHTER_SS_CLINCH_ATTEMPTED" : red_fighter_ss_clinch_attempted,
                  "RED_FIGHTER_SS_GROUND_LANDED" : red_fighter_ss_ground_landed,
                  "RED_FIGHTER_SS_GROUND_ATTEMPTED" : red_fighter_ss_ground_attempted,

                  "BLUE_FIGHTER_SS2_LANDED" : blue_fighter_ss2_landed,
                  "BLUE_FIGHTER_SS2_ATTEMPTED" : blue_fighter_ss2_attempted,
                  "BLUE_FIGHTER_SS2_PERCENT" : blue_fighter_ss2_percent,
                  "BLUE_FIGHTER_SS_HEAD_LANDED" : blue_fighter_ss_head_landed,
                  "BLUE_FIGHTER_SS_HEAD_ATTEMPTED" : blue_fighter_ss_head_attempted,
                  "BLUE_FIGHTER_SS_BODY_LANDED" : blue_fighter_ss_body_landed,
                  "BLUE_FIGHTER_SS_BODY_ATTEMPTED" : blue_fighter_ss_body_attempted,
                  "BLUE_FIGHTER_SS_LEG_LANDED" : blue_fighter_ss_leg_landed,
                  "BLUE_FIGHTER_SS_LEG_ATTEMPTED" : blue_fighter_ss_leg_attempted,
                  "BLUE_FIGHTER_SS_DISTANCE_LANDED" : blue_fighter_ss_distance_landed,
                  "BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED" : blue_fighter_ss_distance_attempted,
                  "BLUE_FIGHTER_SS_CLINCH_LANDED" : blue_fighter_ss_clinch_landed,
                  "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED" : blue_fighter_ss_clinch_attempted,
                  "BLUE_FIGHTER_SS_GROUND_LANDED" : blue_fighter_ss_ground_landed,
                  "BLUE_FIGHTER_SS_GROUND_ATTEMPTED" : blue_fighter_ss_ground_attempted
                  })


        # SIGNIFICANT STRIKES TOTALS ROUND BY ROUND

    for i in range(2, len(RBR_elements2), 2):

        round = ("_R" + str(int(i / 2)))
        #print(round)

        # Significant Strikes

        fighter_ss2_count = RBR_elements2[i].text
        red_fighter_ss2_landed_r = fighter_ss2_count.split("\n")[2].split(" of ")[0]
        red_fighter_ss2_attempted_r = fighter_ss2_count.split("\n")[2].split(" of ")[1]
        blue_fighter_ss2_landed_r = fighter_ss2_count.split("\n")[3].split(" of ")[0]
        blue_fighter_ss2_attempted_r = fighter_ss2_count.split("\n")[3].split(" of ")[1]


        # Significant Strikes Percentage

        fighter_ss2_percent = RBR_elements2[i].text
        red_fighter_ss2_percent_r = fighter_ss2_percent.split("\n")[4]
        blue_fighter_ss2_percent_r = fighter_ss2_percent.split("\n")[5]


        # Head

        fighter_ss_head_count = RBR_elements2[i].text
        red_fighter_ss_head_landed_r = fighter_ss_head_count.split("\n")[6].split(" of ")[0]
        red_fighter_ss_head_attempted_r = fighter_ss_head_count.split("\n")[6].split(" of ")[1]
        blue_fighter_ss_head_landed_r = fighter_ss_head_count.split("\n")[7].split(" of ")[0]
        blue_fighter_ss_head_attempted_r = fighter_ss_head_count.split("\n")[7].split(" of ")[1]


        # Body

        fighter_ss_body_count = RBR_elements2[i].text
        red_fighter_ss_body_landed_r = fighter_ss_body_count.split("\n")[8].split(" of ")[0]
        red_fighter_ss_body_attempted_r = fighter_ss_body_count.split("\n")[8].split(" of ")[1]
        blue_fighter_ss_body_landed_r = fighter_ss_body_count.split("\n")[9].split(" of ")[0]
        blue_fighter_ss_body_attempted_r = fighter_ss_body_count.split("\n")[9].split(" of ")[1]


        # Leg

        fighter_ss_leg_count = RBR_elements2[i].text
        red_fighter_ss_leg_landed_r = fighter_ss_leg_count.split("\n")[10].split(" of ")[0]
        red_fighter_ss_leg_attempted_r = fighter_ss_leg_count.split("\n")[10].split(" of ")[1]
        blue_fighter_ss_leg_landed_r = fighter_ss_leg_count.split("\n")[11].split(" of ")[0]
        blue_fighter_ss_leg_attempted_r = fighter_ss_leg_count.split("\n")[11].split(" of ")[1]


        # Distance

        fighter_ss_distance_count = RBR_elements2[i].text
        red_fighter_ss_distance_landed_r = fighter_ss_distance_count.split("\n")[12].split(" of ")[0]
        red_fighter_ss_distance_attempted_r = fighter_ss_distance_count.split("\n")[12].split(" of ")[1]
        blue_fighter_ss_distance_landed_r = fighter_ss_distance_count.split("\n")[13].split(" of ")[0]
        blue_fighter_ss_distance_attempted_r = fighter_ss_distance_count.split("\n")[13].split(" of ")[1]


        # Clinch

        fighter_ss_clinch_count = RBR_elements2[i].text
        red_fighter_ss_clinch_landed_r = fighter_ss_clinch_count.split("\n")[14].split(" of ")[0]
        red_fighter_ss_clinch_attempted_r = fighter_ss_clinch_count.split("\n")[14].split(" of ")[1]
        blue_fighter_ss_clinch_landed_r = fighter_ss_clinch_count.split("\n")[15].split(" of ")[0]
        blue_fighter_ss_clinch_attempted_r = fighter_ss_clinch_count.split("\n")[15].split(" of ")[1]


        # Ground

        fighter_ss_ground_count = RBR_elements2[i].text
        red_fighter_ss_ground_landed_r = fighter_ss_ground_count.split("\n")[16].split(" of ")[0]
        red_fighter_ss_ground_attempted_r = fighter_ss_ground_count.split("\n")[16].split(" of ")[1]
        blue_fighter_ss_ground_landed_r = fighter_ss_ground_count.split("\n")[17].split(" of ")[0]
        blue_fighter_ss_ground_attempted_r = fighter_ss_ground_count.split("\n")[17].split(" of ")[1]

        stats.append({"RED_FIGHTER_SS2_LANDED"+round : red_fighter_ss2_landed_r,
                      "RED_FIGHTER_SS2_ATTEMPTED"+round : red_fighter_ss2_attempted_r,
                      "RED_FIGHTER_SS2_PERCENT"+round : red_fighter_ss2_percent_r,
                      "RED_FIGHTER_SS_HEAD_LANDED"+round : red_fighter_ss_head_landed_r,
                      "RED_FIGHTER_SS_HEAD_ATTEMPTED"+round : red_fighter_ss_head_attempted_r,
                      "RED_FIGHTER_SS_BODY_LANDED"+round : red_fighter_ss_body_landed_r,
                      "RED_FIGHTER_SS_BODY_ATTEMPTED"+round : red_fighter_ss_body_attempted_r,
                      "RED_FIGHTER_SS_LEG_LANDED"+round : red_fighter_ss_leg_landed_r,
                      "RED_FIGHTER_SS_LEG_ATTEMPTED"+round : red_fighter_ss_leg_attempted_r,
                      "RED_FIGHTER_SS_DISTANCE_LANDED"+round : red_fighter_ss_distance_landed_r,
                      "RED_FIGHTER_SS_DISTANCE_ATTEMPTED"+round : red_fighter_ss_distance_attempted_r,
                      "RED_FIGHTER_SS_CLINCH_LANDED"+round : red_fighter_ss_clinch_landed_r,
                      "RED_FIGHTER_SS_CLINCH_ATTEMPTED"+round : red_fighter_ss_clinch_attempted_r,
                      "RED_FIGHTER_SS_GROUND_LANDED"+round : red_fighter_ss_ground_landed_r,
                      "RED_FIGHTER_SS_GROUND_ATTEMPTED"+round : red_fighter_ss_ground_attempted_r,

                      "BLUE_FIGHTER_SS2_LANDED"+round: blue_fighter_ss2_landed_r,
                      "BLUE_FIGHTER_SS2_ATTEMPTED"+round: blue_fighter_ss2_attempted_r,
                      "BLUE_FIGHTER_SS2_PERCENT"+round: blue_fighter_ss2_percent_r,
                      "BLUE_FIGHTER_SS_HEAD_LANDED"+round: blue_fighter_ss_head_landed_r,
                      "BLUE_FIGHTER_SS_HEAD_ATTEMPTED"+round: blue_fighter_ss_head_attempted_r,
                      "BLUE_FIGHTER_SS_BODY_LANDED"+round: blue_fighter_ss_body_landed_r,
                      "BLUE_FIGHTER_SS_BODY_ATTEMPTED"+round: blue_fighter_ss_body_attempted_r,
                      "BLUE_FIGHTER_SS_LEG_LANDED"+round: blue_fighter_ss_leg_landed_r,
                      "BLUE_FIGHTER_SS_LEG_ATTEMPTED"+round: blue_fighter_ss_leg_attempted_r,
                      "BLUE_FIGHTER_SS_DISTANCE_LANDED"+round: blue_fighter_ss_distance_landed_r,
                      "BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED"+round: blue_fighter_ss_distance_attempted_r,
                      "BLUE_FIGHTER_SS_CLINCH_LANDED"+round: blue_fighter_ss_clinch_landed_r,
                      "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED"+round: blue_fighter_ss_clinch_attempted_r,
                      "BLUE_FIGHTER_SS_GROUND_LANDED"+round: blue_fighter_ss_ground_landed_r,
                      "BLUE_FIGHTER_SS_GROUND_ATTEMPTED"+round: blue_fighter_ss_ground_attempted_r
                      })

        # counter = 0
        # for stat in stats:
        #     for key, value in stat.items():
        #         print(key, ":", value)
        #         counter += 1
        #         if counter % 15 == 0:
        #             print("\n")


    column_names = ["EVENT", "BOUT", "DIVISION", "ROUNDS", "WINNER", "METHOD OF WIN", "FINAL ROUND", "TIME", "RED_FIGHTER", "BLUE_FIGHTER", "RED_FIGHTER_KD", "RED_FIGHTER_SS_LANDED", "RED_FIGHTER_SS_ATTEMPTED", "RED_FIGHTER_SS_PERCENT", "RED_FIGHTER_TS_LANDED", "RED_FIGHTER_TS_ATTEMPTED", "RED_FIGHTER_TD_LANDED", "RED_FIGHTER_TD_ATTEMPTED", "RED_FIGHTER_TD_PERCENT", "RED_FIGHTER_SUB_ATTEMPTS", "RED_FIGHTER_REVERSALS", "RED_FIGHTER_CONTROL_TIME", "BLUE_FIGHTER_KD", "BLUE_FIGHTER_SS_LANDED", "BLUE_FIGHTER_SS_ATTEMPTED", "BLUE_FIGHTER_SS_PERCENT", "BLUE_FIGHTER_TS_LANDED", "BLUE_FIGHTER_TS_ATTEMPTED", "BLUE_FIGHTER_TD_LANDED", "BLUE_FIGHTER_TD_ATTEMPTED", "BLUE_FIGHTER_TD_PERCENT",
                    "BLUE_FIGHTER_SUB_ATTEMPTS", "BLUE_FIGHTER_REVERSALS", "BLUE_FIGHTER_CONTROL_TIME", "RED_FIGHTER_KD_R1", "RED_FIGHTER_SS_LANDED_R1", "RED_FIGHTER_SS_ATTEMPTED_R1", "RED_FIGHTER_SS_PERCENT_R1", "RED_FIGHTER_TS_LANDED_R1", "RED_FIGHTER_TS_ATTEMPTED_R1", "RED_FIGHTER_TD_LANDED_R1", "RED_FIGHTER_TD_ATTEMPTED_R1", "RED_FIGHTER_TD_PERCENT_R1", "RED_FIGHTER_SUB_ATTEMPTS_R1", "RED_FIGHTER_REVERSALS_R1", "RED_FIGHTER_CONTROL_TIME_R1", "BLUE_FIGHTER_KD_R1", "BLUE_FIGHTER_SS_LANDED_R1", "BLUE_FIGHTER_SS_ATTEMPTED_R1", "BLUE_FIGHTER_SS_PERCENT_R1", "BLUE_FIGHTER_TS_LANDED_R1", "BLUE_FIGHTER_TS_ATTEMPTED_R1", "BLUE_FIGHTER_TD_LANDED_R1", "BLUE_FIGHTER_TD_ATTEMPTED_R1", "BLUE_FIGHTER_TD_PERCENT_R1", "BLUE_FIGHTER_SUB_ATTEMPTS_R1", "BLUE_FIGHTER_REVERSALS_R1", "BLUE_FIGHTER_CONTROL_TIME_R1", "RED_FIGHTER_KD_R2", "RED_FIGHTER_SS_LANDED_R2", "RED_FIGHTER_SS_ATTEMPTED_R2",
                    "RED_FIGHTER_SS_PERCENT_R2", "RED_FIGHTER_TS_LANDED_R2", "RED_FIGHTER_TS_ATTEMPTED_R2", "RED_FIGHTER_TD_LANDED_R2", "RED_FIGHTER_TD_ATTEMPTED_R2", "RED_FIGHTER_TD_PERCENT_R2", "RED_FIGHTER_SUB_ATTEMPTS_R2", "RED_FIGHTER_REVERSALS_R2", "RED_FIGHTER_CONTROL_TIME_R2", "BLUE_FIGHTER_KD_R2", "BLUE_FIGHTER_SS_LANDED_R2", "BLUE_FIGHTER_SS_ATTEMPTED_R2", "BLUE_FIGHTER_SS_PERCENT_R2", "BLUE_FIGHTER_TS_LANDED_R2", "BLUE_FIGHTER_TS_ATTEMPTED_R2", "BLUE_FIGHTER_TD_LANDED_R2", "BLUE_FIGHTER_TD_ATTEMPTED_R2", "BLUE_FIGHTER_TD_PERCENT_R2", "BLUE_FIGHTER_SUB_ATTEMPTS_R2", "BLUE_FIGHTER_REVERSALS_R2", "BLUE_FIGHTER_CONTROL_TIME_R2", "RED_FIGHTER_KD_R3", "RED_FIGHTER_SS_LANDED_R3", "RED_FIGHTER_SS_ATTEMPTED_R3", "RED_FIGHTER_SS_PERCENT_R3", "RED_FIGHTER_TS_LANDED_R3", "RED_FIGHTER_TS_ATTEMPTED_R3", "RED_FIGHTER_TD_LANDED_R3", "RED_FIGHTER_TD_ATTEMPTED_R3", "RED_FIGHTER_TD_PERCENT_R3",
                    "RED_FIGHTER_SUB_ATTEMPTS_R3", "RED_FIGHTER_REVERSALS_R3", "RED_FIGHTER_CONTROL_TIME_R3", "BLUE_FIGHTER_KD_R3", "BLUE_FIGHTER_SS_LANDED_R3", "BLUE_FIGHTER_SS_ATTEMPTED_R3", "BLUE_FIGHTER_SS_PERCENT_R3", "BLUE_FIGHTER_TS_LANDED_R3", "BLUE_FIGHTER_TS_ATTEMPTED_R3", "BLUE_FIGHTER_TD_LANDED_R3", "BLUE_FIGHTER_TD_ATTEMPTED_R3", "BLUE_FIGHTER_TD_PERCENT_R3", "BLUE_FIGHTER_SUB_ATTEMPTS_R3", "BLUE_FIGHTER_REVERSALS_R3", "BLUE_FIGHTER_CONTROL_TIME_R3", "RED_FIGHTER_KD_R4", "RED_FIGHTER_SS_LANDED_R4", "RED_FIGHTER_SS_ATTEMPTED_R4", "RED_FIGHTER_SS_PERCENT_R4", "RED_FIGHTER_TS_LANDED_R4", "RED_FIGHTER_TS_ATTEMPTED_R4", "RED_FIGHTER_TD_LANDED_R4", "RED_FIGHTER_TD_ATTEMPTED_R4", "RED_FIGHTER_TD_PERCENT_R4", "RED_FIGHTER_SUB_ATTEMPTS_R4", "RED_FIGHTER_REVERSALS_R4", "RED_FIGHTER_CONTROL_TIME_R4", "BLUE_FIGHTER_KD_R4", "BLUE_FIGHTER_SS_LANDED_R4", "BLUE_FIGHTER_SS_ATTEMPTED_R4",
                    "BLUE_FIGHTER_SS_PERCENT_R4","BLUE_FIGHTER_TS_LANDED_R4","BLUE_FIGHTER_TS_ATTEMPTED_R4","BLUE_FIGHTER_TD_LANDED_R4","BLUE_FIGHTER_TD_ATTEMPTED_R4","BLUE_FIGHTER_TD_PERCENT_R4","BLUE_FIGHTER_SUB_ATTEMPTS_R4","BLUE_FIGHTER_REVERSALS_R4","BLUE_FIGHTER_CONTROL_TIME_R4","RED_FIGHTER_KD_R5","RED_FIGHTER_SS_LANDED_R5","RED_FIGHTER_SS_ATTEMPTED_R5","RED_FIGHTER_SS_PERCENT_R5","RED_FIGHTER_TS_LANDED_R5","RED_FIGHTER_TS_ATTEMPTED_R5","RED_FIGHTER_TD_LANDED_R5","RED_FIGHTER_TD_ATTEMPTED_R5","RED_FIGHTER_TD_PERCENT_R5","RED_FIGHTER_SUB_ATTEMPTS_R5","RED_FIGHTER_REVERSALS_R5","RED_FIGHTER_CONTROL_TIME_R5","BLUE_FIGHTER_KD_R5","BLUE_FIGHTER_SS_LANDED_R5","BLUE_FIGHTER_SS_ATTEMPTED_R5","BLUE_FIGHTER_SS_PERCENT_R5","BLUE_FIGHTER_TS_LANDED_R5","BLUE_FIGHTER_TS_ATTEMPTED_R5","BLUE_FIGHTER_TD_LANDED_R5","BLUE_FIGHTER_TD_ATTEMPTED_R5","BLUE_FIGHTER_TD_PERCENT_R5",
                    "BLUE_FIGHTER_SUB_ATTEMPTS_R5","BLUE_FIGHTER_REVERSALS_R5","BLUE_FIGHTER_CONTROL_TIME_R5","RED_FIGHTER_SS2_LANDED","RED_FIGHTER_SS2_ATTEMPTED","RED_FIGHTER_SS2_PERCENT","RED_FIGHTER_SS_HEAD_LANDED","RED_FIGHTER_SS_HEAD_ATTEMPTED","RED_FIGHTER_SS_BODY_LANDED","RED_FIGHTER_SS_BODY_ATTEMPTED","RED_FIGHTER_SS_LEG_LANDED","RED_FIGHTER_SS_LEG_ATTEMPTED","RED_FIGHTER_SS_DISTANCE_LANDED","RED_FIGHTER_SS_DISTANCE_ATTEMPTED","RED_FIGHTER_SS_CLINCH_LANDED","RED_FIGHTER_SS_CLINCH_ATTEMPTED","RED_FIGHTER_SS_GROUND_LANDED","RED_FIGHTER_SS_GROUND_ATTEMPTED","BLUE_FIGHTER_SS2_LANDED","BLUE_FIGHTER_SS2_ATTEMPTED","BLUE_FIGHTER_SS2_PERCENT","BLUE_FIGHTER_SS_HEAD_LANDED","BLUE_FIGHTER_SS_HEAD_ATTEMPTED","BLUE_FIGHTER_SS_BODY_LANDED","BLUE_FIGHTER_SS_BODY_ATTEMPTED","BLUE_FIGHTER_SS_LEG_LANDED","BLUE_FIGHTER_SS_LEG_ATTEMPTED","BLUE_FIGHTER_SS_DISTANCE_LANDED","BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED","BLUE_FIGHTER_SS_CLINCH_LANDED",
                    "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED", "BLUE_FIGHTER_SS_GROUND_LANDED", "BLUE_FIGHTER_SS_GROUND_ATTEMPTED", "RED_FIGHTER_SS2_LANDED_R1", "RED_FIGHTER_SS2_ATTEMPTED_R1", "RED_FIGHTER_SS2_PERCENT_R1", "RED_FIGHTER_SS_HEAD_LANDED_R1", "RED_FIGHTER_SS_HEAD_ATTEMPTED_R1", "RED_FIGHTER_SS_BODY_LANDED_R1", "RED_FIGHTER_SS_BODY_ATTEMPTED_R1", "RED_FIGHTER_SS_LEG_LANDED_R1", "RED_FIGHTER_SS_LEG_ATTEMPTED_R1", "RED_FIGHTER_SS_DISTANCE_LANDED_R1", "RED_FIGHTER_SS_DISTANCE_ATTEMPTED_R1", "RED_FIGHTER_SS_CLINCH_LANDED_R1", "RED_FIGHTER_SS_CLINCH_ATTEMPTED_R1", "RED_FIGHTER_SS_GROUND_LANDED_R1", "RED_FIGHTER_SS_GROUND_ATTEMPTED_R1", "BLUE_FIGHTER_SS2_LANDED_R1", "BLUE_FIGHTER_SS2_ATTEMPTED_R1", "BLUE_FIGHTER_SS2_PERCENT_R1", "BLUE_FIGHTER_SS_HEAD_LANDED_R1", "BLUE_FIGHTER_SS_HEAD_ATTEMPTED_R1", "BLUE_FIGHTER_SS_BODY_LANDED_R1", "BLUE_FIGHTER_SS_BODY_ATTEMPTED_R1", "BLUE_FIGHTER_SS_LEG_LANDED_R1", "BLUE_FIGHTER_SS_LEG_ATTEMPTED_R1", "BLUE_FIGHTER_SS_DISTANCE_LANDED_R1", "BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED_R1", "BLUE_FIGHTER_SS_CLINCH_LANDED_R1",
                    "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED_R1","BLUE_FIGHTER_SS_GROUND_LANDED_R1","BLUE_FIGHTER_SS_GROUND_ATTEMPTED_R1","RED_FIGHTER_SS2_LANDED_R2","RED_FIGHTER_SS2_ATTEMPTED_R2","RED_FIGHTER_SS2_PERCENT_R2","RED_FIGHTER_SS_HEAD_LANDED_R2","RED_FIGHTER_SS_HEAD_ATTEMPTED_R2","RED_FIGHTER_SS_BODY_LANDED_R2","RED_FIGHTER_SS_BODY_ATTEMPTED_R2","RED_FIGHTER_SS_LEG_LANDED_R2","RED_FIGHTER_SS_LEG_ATTEMPTED_R2","RED_FIGHTER_SS_DISTANCE_LANDED_R2","RED_FIGHTER_SS_DISTANCE_ATTEMPTED_R2","RED_FIGHTER_SS_CLINCH_LANDED_R2","RED_FIGHTER_SS_CLINCH_ATTEMPTED_R2","RED_FIGHTER_SS_GROUND_LANDED_R2","RED_FIGHTER_SS_GROUND_ATTEMPTED_R2","BLUE_FIGHTER_SS2_LANDED_R2","BLUE_FIGHTER_SS2_ATTEMPTED_R2","BLUE_FIGHTER_SS2_PERCENT_R2","BLUE_FIGHTER_SS_HEAD_LANDED_R2","BLUE_FIGHTER_SS_HEAD_ATTEMPTED_R2","BLUE_FIGHTER_SS_BODY_LANDED_R2","BLUE_FIGHTER_SS_BODY_ATTEMPTED_R2","BLUE_FIGHTER_SS_LEG_LANDED_R2","BLUE_FIGHTER_SS_LEG_ATTEMPTED_R2","BLUE_FIGHTER_SS_DISTANCE_LANDED_R2","BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED_R2","BLUE_FIGHTER_SS_CLINCH_LANDED_R2",
                    "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED_R2","BLUE_FIGHTER_SS_GROUND_LANDED_R2","BLUE_FIGHTER_SS_GROUND_ATTEMPTED_R2","RED_FIGHTER_SS2_LANDED_R3","RED_FIGHTER_SS2_ATTEMPTED_R3","RED_FIGHTER_SS2_PERCENT_R3","RED_FIGHTER_SS_HEAD_LANDED_R3","RED_FIGHTER_SS_HEAD_ATTEMPTED_R3","RED_FIGHTER_SS_BODY_LANDED_R3","RED_FIGHTER_SS_BODY_ATTEMPTED_R3","RED_FIGHTER_SS_LEG_LANDED_R3","RED_FIGHTER_SS_LEG_ATTEMPTED_R3","RED_FIGHTER_SS_DISTANCE_LANDED_R3","RED_FIGHTER_SS_DISTANCE_ATTEMPTED_R3","RED_FIGHTER_SS_CLINCH_LANDED_R3","RED_FIGHTER_SS_CLINCH_ATTEMPTED_R3","RED_FIGHTER_SS_GROUND_LANDED_R3","RED_FIGHTER_SS_GROUND_ATTEMPTED_R3","BLUE_FIGHTER_SS2_LANDED_R3","BLUE_FIGHTER_SS2_ATTEMPTED_R3","BLUE_FIGHTER_SS2_PERCENT_R3","BLUE_FIGHTER_SS_HEAD_LANDED_R3","BLUE_FIGHTER_SS_HEAD_ATTEMPTED_R3","BLUE_FIGHTER_SS_BODY_LANDED_R3","BLUE_FIGHTER_SS_BODY_ATTEMPTED_R3","BLUE_FIGHTER_SS_LEG_LANDED_R3","BLUE_FIGHTER_SS_LEG_ATTEMPTED_R3","BLUE_FIGHTER_SS_DISTANCE_LANDED_R3","BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED_R3","BLUE_FIGHTER_SS_CLINCH_LANDED_R3",
                    "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED_R3", "BLUE_FIGHTER_SS_GROUND_LANDED_R3", "BLUE_FIGHTER_SS_GROUND_ATTEMPTED_R3", "RED_FIGHTER_SS2_LANDED_R4", "RED_FIGHTER_SS2_ATTEMPTED_R4", "RED_FIGHTER_SS2_PERCENT_R4", "RED_FIGHTER_SS_HEAD_LANDED_R4", "RED_FIGHTER_SS_HEAD_ATTEMPTED_R4", "RED_FIGHTER_SS_BODY_LANDED_R4", "RED_FIGHTER_SS_BODY_ATTEMPTED_R4", "RED_FIGHTER_SS_LEG_LANDED_R4", "RED_FIGHTER_SS_LEG_ATTEMPTED_R4", "RED_FIGHTER_SS_DISTANCE_LANDED_R4", "RED_FIGHTER_SS_DISTANCE_ATTEMPTED_R4", "RED_FIGHTER_SS_CLINCH_LANDED_R4", "RED_FIGHTER_SS_CLINCH_ATTEMPTED_R4", "RED_FIGHTER_SS_GROUND_LANDED_R4", "RED_FIGHTER_SS_GROUND_ATTEMPTED_R4", "BLUE_FIGHTER_SS2_LANDED_R4", "BLUE_FIGHTER_SS2_ATTEMPTED_R4", "BLUE_FIGHTER_SS2_PERCENT_R4", "BLUE_FIGHTER_SS_HEAD_LANDED_R4", "BLUE_FIGHTER_SS_HEAD_ATTEMPTED_R4", "BLUE_FIGHTER_SS_BODY_LANDED_R4", "BLUE_FIGHTER_SS_BODY_ATTEMPTED_R4", "BLUE_FIGHTER_SS_LEG_LANDED_R4", "BLUE_FIGHTER_SS_LEG_ATTEMPTED_R4", "BLUE_FIGHTER_SS_DISTANCE_LANDED_R4", "BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED_R4", "BLUE_FIGHTER_SS_CLINCH_LANDED_R4",
                    "BLUE_FIGHTER_SS_CLINCH_ATTEMPTED_R4","BLUE_FIGHTER_SS_GROUND_LANDED_R4","BLUE_FIGHTER_SS_GROUND_ATTEMPTED_R4","RED_FIGHTER_SS2_LANDED_R5","RED_FIGHTER_SS2_ATTEMPTED_R5","RED_FIGHTER_SS2_PERCENT_R5","RED_FIGHTER_SS_HEAD_LANDED_R5","RED_FIGHTER_SS_HEAD_ATTEMPTED_R5","RED_FIGHTER_SS_BODY_LANDED_R5","RED_FIGHTER_SS_BODY_ATTEMPTED_R5","RED_FIGHTER_SS_LEG_LANDED_R5","RED_FIGHTER_SS_LEG_ATTEMPTED_R5","RED_FIGHTER_SS_DISTANCE_LANDED_R5","RED_FIGHTER_SS_DISTANCE_ATTEMPTED_R5","RED_FIGHTER_SS_CLINCH_LANDED_R5","RED_FIGHTER_SS_CLINCH_ATTEMPTED_R5","RED_FIGHTER_SS_GROUND_LANDED_R5","RED_FIGHTER_SS_GROUND_ATTEMPTED_R5","BLUE_FIGHTER_SS2_LANDED_R5","BLUE_FIGHTER_SS2_ATTEMPTED_R5","BLUE_FIGHTER_SS2_PERCENT_R5","BLUE_FIGHTER_SS_HEAD_LANDED_R5","BLUE_FIGHTER_SS_HEAD_ATTEMPTED_R5","BLUE_FIGHTER_SS_BODY_LANDED_R5","BLUE_FIGHTER_SS_BODY_ATTEMPTED_R5","BLUE_FIGHTER_SS_LEG_LANDED_R5","BLUE_FIGHTER_SS_LEG_ATTEMPTED_R5","BLUE_FIGHTER_SS_DISTANCE_LANDED_R5","BLUE_FIGHTER_SS_DISTANCE_ATTEMPTED_R5","BLUE_FIGHTER_SS_CLINCH_LANDED_R5","BLUE_FIGHTER_SS_CLINCH_ATTEMPTED_R5","BLUE_FIGHTER_SS_GROUND_LANDED_R5","BLUE_FIGHTER_SS_GROUND_ATTEMPTED_R5"]

    # Create an empty dictionary with keys from the list of column names, should replace N/A with ---
    result_dict = {col: "---" for col in column_names}

    if 'Sheet1' in book.sheetnames:

        merged_dict = {k: v for d in stats for k, v in d.items()}
        # stats2 = [merged_dict]
        #
        # df = pd.DataFrame(stats2)
        # Update the empty dictionary with the values from the data_dict
        result_dict.update(merged_dict)
        # Convert the dictionary to a dataframe
        df = pd.DataFrame(result_dict, index=[s])
        final_df = final_df.append(df, ignore_index=True)
        #print(result_dict)

        s = s+1

        #print(df,"\n")
        # Fill in any missing columns from the list of column names with "NA"
        #df = df.reindex(columns=column_names)
        #print(df, "\n")
        #df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=s)

    else:

        print("Sheet Doesnt Exist")

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
#print(final_df)

for i in range(len(final_df.index)):
    for j in range(len(final_df.columns)):
        # check if the cell consists only of a number
        if str(final_df.iat[i,j]).isdigit():
            # convert the cell to a numeric value
            final_df.iat[i,j] = pd.to_numeric(final_df.iat[i,j])

final_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=1)
writer.close()